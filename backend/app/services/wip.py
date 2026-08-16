"""Work-in-progress reporting and escalation.

Three WIP views (locked decisions):
  * Today  — today's activity, grouped by person (technicians, helpers, engineers, admin)
  * Past   — completed work history over a date range
  * Future — scheduled work with dates (PMS visits + tasks due later)

Escalation is driven by INACTIVITY, not ticket age: a ticket being worked on daily never
escalates, while a forgotten one surfaces quickly.
  * no activity for ESCALATION_L1_DAYS -> Service Engineer + Service Admin
  * no activity for ESCALATION_L2_DAYS -> Managing Director
An escalation clears automatically once the ticket closes (closed tickets are never scanned).
"""

from datetime import date, timedelta

from sqlmodel import Session, select

from app.core.enums import (
    ESCALATION_L1_DAYS,
    ESCALATION_L2_DAYS,
    LifecycleStage,
    TaskStatus,
    TicketStatus,
    WorkType,
)
from app.models.masters import Customer, TeamMember
from app.models.pms import PMS, PMSVisitTicket
from app.models.task import Task
from app.models.tickets import Ticket, TicketUpdate
from app.models.user import User
from app.services.pms_schedule import generate_visit_dates

# Statuses that count as unfinished work.
OPEN_STATUSES = (TicketStatus.OPEN, TicketStatus.IN_PROGRESS, TicketStatus.REOPENED)

# Bucket for activity recorded without a job lead or team member.
UNASSIGNED = "Unassigned"

# Stages that mean physical work is under way (used to find when a job actually started).
_WORK_STAGES = {
    LifecycleStage.WORK_STARTED,
    LifecycleStage.MATERIAL_PENDING,
    LifecycleStage.TESTING_COMMISSIONING,
    # deprecated equivalents, so historical tickets still resolve
    LifecycleStage.DIAGNOSED,
    LifecycleStage.PARTS_REQUESTED,
    LifecycleStage.REPAIR_IN_PROGRESS,
}


def work_started_date(ticket: Ticket) -> date | None:
    """Earliest date real work began (a Work Started or later stage). None if not started."""
    dates: list[date] = []
    for u in ticket.updates:
        if u.stage in _WORK_STAGES:
            for d in (u.start_date, u.action_date):
                if d is not None:
                    dates.append(d)
    return min(dates) if dates else None


def closed_date(ticket: Ticket) -> date | None:
    """The date the ticket was closed (latest end date), or None if still open."""
    ends = [u.end_date for u in ticket.updates if u.end_date is not None]
    return max(ends) if ends else None


def active_on(ticket: Ticket, day: date) -> bool:
    """True if the ticket's work was in progress on `day` — started on/before it and not yet
    closed before it. This is what makes a multi-day job appear on every day it spans, even
    when no fresh lifecycle row was logged that day."""
    started = work_started_date(ticket)
    if started is None or started > day:
        return False
    closed = closed_date(ticket)
    return closed is None or closed >= day


def _people_asof(ticket: Ticket, day: date) -> list[str]:
    """The crew on the ticket as known on `day`: job lead + team from the most recent
    lifecycle row dated on/before that day. Falls back to any people ever recorded."""
    def _dated(u) -> date | None:
        ds = [d for d in (u.action_date, u.start_date, u.end_date) if d is not None]
        return max(ds) if ds else None

    candidates = [u for u in ticket.updates if (_dated(u) or day) <= day]
    for u in sorted(candidates or ticket.updates, key=lambda x: x.id or 0, reverse=True):
        people = [n for n in ([u.job_lead] + [m.name for m in u.team]) if n]
        if people:
            return people
    return []


def last_activity_date(ticket: Ticket) -> date:
    """Most recent date anything was recorded on the ticket.

    Falls back to the complaint date when no lifecycle row carries a date, so a freshly
    logged ticket starts its inactivity clock from the day it came in.
    """
    dates = [
        d
        for u in ticket.updates
        for d in (u.action_date, u.start_date, u.end_date)
        if d is not None
    ]
    return max(dates) if dates else ticket.complaint_date


def escalation_level(ticket: Ticket, today: date | None = None) -> int:
    """0 = fine, 1 = Engineer/Admin, 2 = Managing Director. Closed tickets are always 0."""
    if ticket.status not in OPEN_STATUSES:
        return 0
    today = today or date.today()
    idle = (today - last_activity_date(ticket)).days
    if idle >= ESCALATION_L2_DAYS:
        return 2
    if idle >= ESCALATION_L1_DAYS:
        return 1
    return 0


def _lead_and_team(ticket: Ticket, day: date) -> tuple[str | None, list[str]]:
    """The job lead and the rest of the crew as known on `day`, taken from the most recent
    lifecycle row dated on/before it (the same source _people_asof uses)."""
    def _dated(u) -> date | None:
        ds = [d for d in (u.action_date, u.start_date, u.end_date) if d is not None]
        return max(ds) if ds else None

    candidates = [u for u in ticket.updates if (_dated(u) or day) <= day]
    for u in sorted(candidates or ticket.updates, key=lambda x: x.id or 0, reverse=True):
        team = [m.name for m in u.team]
        if u.job_lead or team:
            others = [n for n in team if n != u.job_lead]
            return u.job_lead, others
    return None, []


def current_stage(ticket: Ticket) -> str:
    """The stage of the most recent lifecycle row — the ticket's current position in the flow."""
    if not ticket.updates:
        return "Logged"
    return max(ticket.updates, key=lambda u: u.id or 0).stage.value


def _ticket_brief(t: Ticket, customers: dict[int, Customer], today: date) -> dict:
    cust = customers.get(t.customer_id)
    lead, team = _lead_and_team(t, today)
    return {
        "id": t.id,
        "ticket_no": t.ticket_no,
        "customer_name": cust.name if cust else None,
        "customer_city": cust.city if cust else None,
        "job_lead": lead,
        "team": team,
        "work_type": t.work_type.value,
        "status": t.status.value,
        "stage": current_stage(t),
        "complaint_date": t.complaint_date.isoformat(),
        "last_activity": last_activity_date(t).isoformat(),
        "idle_days": (today - last_activity_date(t)).days,
        "escalation_level": escalation_level(t, today),
    }


def escalations(session: Session, today: date | None = None) -> dict:
    """Open tickets that have gone quiet, split by escalation level."""
    today = today or date.today()
    customers = {c.id: c for c in session.exec(select(Customer)).all()}
    tickets = session.exec(select(Ticket).where(Ticket.status.in_(OPEN_STATUSES))).all()

    level_1, level_2 = [], []
    for t in tickets:
        lvl = escalation_level(t, today)
        if lvl == 2:
            level_2.append(_ticket_brief(t, customers, today))
        elif lvl == 1:
            level_1.append(_ticket_brief(t, customers, today))

    level_1.sort(key=lambda r: r["idle_days"], reverse=True)
    level_2.sort(key=lambda r: r["idle_days"], reverse=True)
    return {
        "as_of": today.isoformat(),
        "l1_days": ESCALATION_L1_DAYS,
        "l2_days": ESCALATION_L2_DAYS,
        "level_1": level_1,   # -> Service Engineer + Service Admin
        "level_2": level_2,   # -> Managing Director
    }


def today_wip(session: Session, today: date | None = None) -> dict:
    """Work-in-progress for a given day, grouped by person, plus each person's open workload.

    A ticket counts for the day if it was either *touched* that day (a lifecycle row is dated
    then) OR its work was *in progress* that day — started on/before it and not yet closed. The
    second rule is what keeps a multi-day job visible on every day it spans, even with no fresh
    update logged. Every roster member is listed regardless, so idle staff stay visible.
    """
    today = today or date.today()
    customers = {c.id: c for c in session.exec(select(Customer)).all()}
    tickets = list(session.exec(select(Ticket)).all())
    members = session.exec(select(TeamMember)).all()

    # name -> {"today": [...], "open": [...]}
    buckets: dict[str, dict[str, list]] = {
        m.name: {"team_type": m.team_type.value, "today": [], "open": []} for m in members
    }

    def _add(name: str | None, key: str, row: dict) -> None:
        if not name:
            return
        bucket = buckets.setdefault(name, {"team_type": "—", "today": [], "open": []})
        if all(r["ticket_no"] != row["ticket_no"] for r in bucket[key]):
            bucket[key].append(row)

    # Flat, customer-centric list of the day's in-progress tickets (one row per ticket) and a
    # count of them by lifecycle stage — both drive the Dashboard's Today-WIP view.
    active_tickets: list[dict] = []
    by_stage: dict[str, int] = {}

    for t in tickets:
        touched = any(
            today in (u.action_date, u.start_date, u.end_date) for u in t.updates
        )
        ongoing = active_on(t, today)
        if touched or ongoing:
            row = _ticket_brief(t, customers, today)
            # Ongoing but not touched today = spanning work carried over from an earlier day.
            started = work_started_date(t)
            row["ongoing"] = ongoing and not touched
            row["started_on"] = started.isoformat() if started else None
            active_tickets.append(row)
            by_stage[row["stage"]] = by_stage.get(row["stage"], 0) + 1
            people = _people_asof(t, today) or [UNASSIGNED]
            for name in people:
                _add(name, "today", row)

        # Open workload is independent of the chosen day — it's "what is still on their plate".
        if t.status in OPEN_STATUSES:
            row_open = _ticket_brief(t, customers, today)
            for name in (_people_asof(t, today) or [UNASSIGNED]):
                _add(name, "open", row_open)

    # Newest complaint first, so the freshest jobs sit at the top of the table.
    active_tickets.sort(key=lambda r: r["complaint_date"], reverse=True)

    people = [
        {
            "name": name,
            "team_type": data["team_type"],
            "today_count": len(data["today"]),
            "open_count": len(data["open"]),
            "today": data["today"],
            "open": data["open"],
        }
        for name, data in buckets.items()
    ]
    # Busiest first; idle staff fall to the bottom but stay visible.
    people.sort(key=lambda p: (-p["today_count"], -p["open_count"], p["name"]))
    return {
        "date": today.isoformat(),
        "active_people": sum(1 for p in people if p["today_count"]),
        "total_touched": len(active_tickets),
        "people": people,           # by-person view (WIP report page)
        "tickets": active_tickets,  # customer-centric flat list (Dashboard table)
        "by_stage": by_stage,       # stage distribution (Dashboard bar chart)
    }


def past_wip(session: Session, start: date, end: date) -> dict:
    """Completed work history: tickets closed within the range."""
    customers = {c.id: c for c in session.exec(select(Customer)).all()}
    rows = []
    for t in session.exec(select(Ticket).where(Ticket.status == TicketStatus.CLOSED)).all():
        ends = [u.end_date for u in t.updates if u.end_date is not None]
        if not ends:
            continue
        closed_on = max(ends)
        if not (start <= closed_on <= end):
            continue
        leads = [u.job_lead for u in t.updates if u.job_lead]
        rows.append({
            "id": t.id,
            "ticket_no": t.ticket_no,
            "customer_name": customers[t.customer_id].name if t.customer_id in customers else None,
            "work_type": t.work_type.value,
            "complaint_date": t.complaint_date.isoformat(),
            "closed_on": closed_on.isoformat(),
            "days_taken": (closed_on - t.complaint_date).days,
            "job_lead": leads[-1] if leads else None,
        })
    rows.sort(key=lambda r: r["closed_on"], reverse=True)
    avg = round(sum(r["days_taken"] for r in rows) / len(rows), 1) if rows else 0.0
    return {
        "start": start.isoformat(),
        "end": end.isoformat(),
        "count": len(rows),
        "avg_days_taken": avg,
        "rows": rows,
    }


def future_wip(session: Session, today: date | None = None) -> dict:
    """Scheduled work ahead: PMS visits not yet ticketed, plus tasks due later."""
    today = today or date.today()
    customers = {c.id: c for c in session.exec(select(Customer)).all()}
    linked = {(lk.pms_id, lk.visit_no) for lk in session.exec(select(PMSVisitTicket)).all()}

    visits = []
    for pms in session.exec(select(PMS)).all():
        for i, d in enumerate(generate_visit_dates(pms.wo_start_date, pms.schedule, pms.wo_end_date), start=1):
            if d > today and (pms.id, i) not in linked:
                visits.append({
                    "pms_id": pms.id,
                    "visit_no": i,
                    "scheduled_on": d.isoformat(),
                    "days_away": (d - today).days,
                    "customer_name": (
                        customers[pms.customer_id].name if pms.customer_id in customers else None
                    ),
                })
    visits.sort(key=lambda r: r["scheduled_on"])

    users = {u.id: u for u in session.exec(select(User)).all()}
    tasks = []
    for t in session.exec(select(Task).where(Task.status != TaskStatus.DONE)).all():
        if t.due_date and t.due_date > today:
            assignee = users.get(t.assignee_user_id)
            tasks.append({
                "id": t.id,
                "title": t.title,
                "due_date": t.due_date.isoformat(),
                "days_away": (t.due_date - today).days,
                "assignee": assignee.full_name or assignee.username if assignee else None,
                "priority": t.priority.value,
            })
    tasks.sort(key=lambda r: r["due_date"])

    return {
        "as_of": today.isoformat(),
        "pms_visits": visits,
        "tasks": tasks,
        "total": len(visits) + len(tasks),
    }


def wip_counts(session: Session, today: date | None = None) -> dict:
    """Headline numbers for the dashboard tiles."""
    today = today or date.today()
    esc = escalations(session, today)
    week_ago = today - timedelta(days=7)
    return {
        "today_touched": today_wip(session, today)["total_touched"],
        "past_7d_closed": past_wip(session, week_ago, today)["count"],
        "future_scheduled": future_wip(session, today)["total"],
        "escalated_l1": len(esc["level_1"]),
        "escalated_l2": len(esc["level_2"]),
    }


def daily_activity(session: Session, start: date, end: date) -> dict:
    """Per-day manpower series over [start, end] for the dashboard's Daily activity charts:

      * closed     — tickets whose latest end date falls on that day (work completed)
      * people     — distinct staff (job lead + team) with any lifecycle row dated that day
      * backlog    — tickets still open at end of that day (logged on/before it, not yet closed)
      * per_person — closed / people (output per head; 0 when nobody was present)
    """
    tickets = list(session.exec(select(Ticket)).all())
    days: list[date] = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)
    day_set = set(days)

    closed = {dd: 0 for dd in days}
    people: dict[date, set[str]] = {dd: set() for dd in days}
    backlog = {dd: 0 for dd in days}

    for t in tickets:
        cd = closed_date(t)
        if cd in closed:
            closed[cd] += 1
        for u in t.updates:
            dated = {x for x in (u.action_date, u.start_date, u.end_date) if x is not None}
            names = [u.job_lead] + [m.name for m in u.team]
            for dd in dated & day_set:
                for n in names:
                    if n:
                        people[dd].add(n)
        for dd in days:
            if t.complaint_date <= dd and (cd is None or cd > dd):
                backlog[dd] += 1

    series = []
    for dd in days:
        ppl = len(people[dd])
        cl = closed[dd]
        series.append({
            "date": dd.isoformat(),
            "closed": cl,
            "people": ppl,
            "backlog": backlog[dd],
            "per_person": round(cl / ppl, 1) if ppl else 0.0,
        })
    return {"start": start.isoformat(), "end": end.isoformat(), "series": series}


def _touched_in_range(ticket: Ticket, start: date, end: date) -> bool:
    for u in ticket.updates:
        for d in (u.action_date, u.start_date, u.end_date):
            if d is not None and start <= d <= end:
                return True
    return False


def _overlaps_range(ticket: Ticket, start: date, end: date) -> bool:
    """Work was in progress at some point within [start, end]."""
    started = work_started_date(ticket)
    if started is None or started > end:
        return False
    closed = closed_date(ticket)
    return closed is None or closed >= start


def period_range(period: str, anchor: date) -> tuple[date, date, str]:
    """Resolve a Daily/Weekly/Monthly preset to a [start, end] range around an anchor date."""
    p = (period or "daily").lower()
    if p == "weekly":
        start = _week_start(anchor)
        return start, start + timedelta(days=6), "Weekly"
    if p == "monthly":
        start = anchor.replace(day=1)
        nxt = (start + timedelta(days=32)).replace(day=1)
        return start, nxt - timedelta(days=1), "Monthly"
    return anchor, anchor, "Daily"


def wip_report(session: Session, start: date, end: date, period_label: str = "Custom") -> dict:
    """On-demand WIP report over a date range: summary counts + ticket list + per-technician.

    A ticket is in the report if it was opened, closed, touched, or in progress within the
    range. Mirrors the WIP page's grouping (per-person workload) but aggregated over the period.
    """
    customers = {c.id: c for c in session.exec(select(Customer)).all()}
    members = session.exec(select(TeamMember)).all()
    today = date.today()

    tickets = list(session.exec(select(Ticket)).all())
    included: list[Ticket] = []
    for t in tickets:
        opened_in = start <= t.complaint_date <= end
        closed_on = closed_date(t)
        closed_in = closed_on is not None and start <= closed_on <= end
        if opened_in or closed_in or _touched_in_range(t, start, end) or _overlaps_range(t, start, end):
            included.append(t)

    # ---- summary counts ----
    by_status: dict[str, int] = {}
    by_work_type: dict[str, int] = {}
    opened = closed = still_open = 0
    for t in included:
        by_status[t.status.value] = by_status.get(t.status.value, 0) + 1
        by_work_type[t.work_type.value] = by_work_type.get(t.work_type.value, 0) + 1
        if start <= t.complaint_date <= end:
            opened += 1
        cd = closed_date(t)
        if cd is not None and start <= cd <= end:
            closed += 1
        if t.status in OPEN_STATUSES:
            still_open += 1

    # ---- ticket list ----
    ticket_rows = [_ticket_brief(t, customers, today) for t in included]
    ticket_rows.sort(key=lambda r: r["complaint_date"], reverse=True)

    # ---- per-technician ----
    def _new_bucket(tt: str = "—") -> dict:
        return {"team_type": tt, "worked": set(), "closed": set(), "open": set(), "days": set()}

    buckets: dict[str, dict] = {m.name: _new_bucket(m.team_type.value) for m in members}

    def _bucket(name: str) -> dict:
        return buckets.setdefault(name, _new_bucket())

    day_range = (start, end)
    for t in included:
        lead, others = _lead_and_team(t, end)
        crew = [n for n in ([lead] + others) if n] or [UNASSIGNED]
        cd = closed_date(t)
        for name in crew:
            b = _bucket(name)
            b["worked"].add(t.ticket_no)
            if t.status in OPEN_STATUSES:
                b["open"].add(t.ticket_no)
        if cd is not None and start <= cd <= end and lead:
            _bucket(lead)["closed"].add(t.ticket_no)
        # Attendance: distinct days each crew member logged activity within the range.
        for u in t.updates:
            dated = [x for x in (u.action_date, u.start_date, u.end_date) if x is not None]
            in_range = [dd for dd in dated if day_range[0] <= dd <= day_range[1]]
            if not in_range:
                continue
            names = [n for n in ([u.job_lead] + [m.name for m in u.team]) if n]
            for name in names:
                for dd in in_range:
                    _bucket(name)["days"].add(dd)

    per_technician = [
        {
            "name": name,
            "team_type": b["team_type"],
            "worked_count": len(b["worked"]),
            "closed_count": len(b["closed"]),
            "open_count": len(b["open"]),
            "active_days": len(b["days"]),
        }
        for name, b in buckets.items()
    ]
    per_technician.sort(key=lambda p: (-p["closed_count"], -p["active_days"], -p["worked_count"], p["name"]))

    # ---- breakdown-call focus (subset of the summary) ----
    bd = [t for t in included if t.work_type == WorkType.BREAKDOWN]
    breakdown = {
        "total": len(bd),
        "opened": sum(1 for t in bd if start <= t.complaint_date <= end),
        "closed": sum(1 for t in bd if (closed_date(t) and start <= closed_date(t) <= end)),
        "still_open": sum(1 for t in bd if t.status in OPEN_STATUSES),
    }

    return {
        "period": period_label,
        "start": start.isoformat(),
        "end": end.isoformat(),
        "summary": {
            "total": len(included),
            "opened": opened,
            "closed": closed,
            "still_open": still_open,
            "by_status": by_status,
            "by_work_type": by_work_type,
            "breakdown": breakdown,
        },
        "tickets": ticket_rows,
        "per_technician": per_technician,
    }


def wip_report_xlsx(report: dict) -> bytes:
    """Render a wip_report() dict to an .xlsx workbook (Summary / Tickets / Per-Technician)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["WIP Report"]); ws["A1"].font = bold
    ws.append(["Period", report["period"]])
    ws.append(["From", report["start"]])
    ws.append(["To", report["end"]])
    ws.append([])
    s = report["summary"]
    ws.append(["Total tickets", s["total"]])
    ws.append(["Opened in period", s["opened"]])
    ws.append(["Closed in period", s["closed"]])
    ws.append(["Still open", s["still_open"]])
    ws.append([])
    ws.append(["By status"]); ws.cell(ws.max_row, 1).font = bold
    for k, v in s["by_status"].items():
        ws.append([k, v])
    ws.append([])
    ws.append(["By work type"]); ws.cell(ws.max_row, 1).font = bold
    for k, v in s["by_work_type"].items():
        ws.append([k, v])

    # Tickets sheet
    wt = wb.create_sheet("Tickets")
    headers = ["Ticket No.", "Customer", "City", "Work Type", "Status", "Stage",
               "Job Lead", "Team", "Complaint Date", "Last Activity", "Idle Days"]
    wt.append(headers)
    for c in range(1, len(headers) + 1):
        wt.cell(1, c).font = bold
    for r in report["tickets"]:
        wt.append([
            r["ticket_no"], r.get("customer_name") or "", r.get("customer_city") or "",
            r["work_type"], r["status"], r["stage"], r.get("job_lead") or "",
            ", ".join(r.get("team") or []), r["complaint_date"], r["last_activity"],
            r["idle_days"],
        ])

    # Per-Technician sheet
    wp = wb.create_sheet("Per-Technician")
    ph = ["Name", "Role", "Worked", "Closed", "Active Days", "Open Workload"]
    wp.append(ph)
    for c in range(1, len(ph) + 1):
        wp.cell(1, c).font = bold
    for p in report["per_technician"]:
        wp.append([p["name"], p["team_type"], p["worked_count"], p["closed_count"],
                   p.get("active_days", 0), p["open_count"]])

    # Auto-ish column widths
    for sheet in (ws, wt, wp):
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(width + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _week_start(d: date) -> date:
    """Monday of the ISO week containing d."""
    return d - timedelta(days=d.weekday())


def backlog_trend(session: Session, weeks: int = 8, today: date | None = None) -> dict:
    """Opened vs closed per week, plus a team attendance grid (present/absent per week).

    * opened  — tickets whose complaint date falls in the week.
    * closed  — tickets whose latest end date falls in the week.
    * present — a team member is 'present' in a week if they were the job lead or on the team
                of any lifecycle row dated within that week; otherwise 'absent'.
    Answers two questions at once: is the backlog growing, and who was actually working.
    """
    today = today or date.today()
    this_week = _week_start(today)
    starts = [this_week - timedelta(weeks=(weeks - 1 - i)) for i in range(weeks)]
    index = {s: i for i, s in enumerate(starts)}

    opened = [0] * weeks
    closed = [0] * weeks
    # week index -> set of names active that week
    active: list[set[str]] = [set() for _ in range(weeks)]

    tickets = session.exec(select(Ticket)).all()
    for t in tickets:
        wk = _week_start(t.complaint_date)
        if wk in index:
            opened[index[wk]] += 1
        ends = [u.end_date for u in t.updates if u.end_date is not None]
        if ends:
            cwk = _week_start(max(ends))
            if cwk in index:
                closed[index[cwk]] += 1
        # attendance from every dated lifecycle row
        for u in t.updates:
            dts = [d for d in (u.action_date, u.start_date, u.end_date) if d is not None]
            if not dts:
                continue
            uwk = _week_start(max(dts))
            if uwk not in index:
                continue
            names = [u.job_lead] + [m.name for m in u.team]
            for n in names:
                if n:
                    active[index[uwk]].add(n)

    members = session.exec(select(TeamMember)).all()
    team = [
        {"name": m.name, "team_type": m.team_type.value,
         "present": [m.name in active[i] for i in range(weeks)]}
        for m in members
    ]
    # Busiest people first; those present in no week fall to the bottom.
    team.sort(key=lambda p: (-sum(p["present"]), p["name"]))

    def _label(s: date) -> str:
        return s.strftime("%d %b")

    return {
        "weeks": [
            {"week_start": s.isoformat(), "label": _label(s),
             "opened": opened[i], "closed": closed[i], "active_count": len(active[i])}
            for i, s in enumerate(starts)
        ],
        "team": team,
    }
