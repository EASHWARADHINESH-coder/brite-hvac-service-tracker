"""Idempotent importers for the monthly PMS / Breakdown Excel sheets.

Two entry points, both safe to re-run (they upsert / skip rather than duplicate):
  * import_pms       — creates Customer master rows from the 'Jul 2026 PMS' sheet
                       (customer details only; no WO / contract / schedule).
  * import_breakdown — creates Breakdown tickets from the 'Breakdown Status' sheet,
                       auto-closing rows that carry a Closing Date.

Dedupe keys: customers by CRM Customer Id (fallback: name); breakdown tickets by
(customer, complaint_date, complaint), and never a second ticket while the customer
already has an open one.
"""

from __future__ import annotations

from datetime import date, datetime

from openpyxl import load_workbook
from sqlalchemy import inspect, text
from sqlmodel import Session, select

from app.core.enums import LifecycleStage, MachineType, TicketStatus, WorkType
from app.models.masters import Complaint, Customer
from app.models.tickets import Ticket, TicketUpdate
from app.services.ticket_logic import compute_ticket_status, derive_skill, next_ticket_no

# ---- Breakdown value mappings (confirmed with the user) ----
COMPLAINT_MAP = {
    "EXV Replace": "EXP Valve Problem",
    "Indoor Motor issue": "IDU Motor Problem",
    "Service Require": "General Service",
    "Cooling issue": "General Complaint",
    "Breakdown": "General Complaint",
    "Not working": "General Complaint",
    # already-known values pass through unchanged
    "Compressor failure": "Compressor failure",
    "Gas Leakage": "Gas Leakage",
}
MACHINE_MAP = {"IVRF": "VRF"}


def _s(v) -> str | None:
    """Stringify a cell value, trimming; blank/None -> None."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _as_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def ensure_schema(engine) -> None:
    """Add the crm_customer_id column to an existing customer table if missing.

    create_all won't ALTER an existing table, so pre-existing dev databases need this.
    ADD COLUMN is supported by both SQLite and Postgres.
    """
    cols = {c["name"] for c in inspect(engine).get_columns("customer")}
    if "crm_customer_id" not in cols:
        with engine.begin() as conn:
            conn.execute(text("ALTER TABLE customer ADD COLUMN crm_customer_id VARCHAR"))


# ---------------------------------------------------------------------------
# PMS -> customers
# ---------------------------------------------------------------------------
def import_pms(session: Session, path: str, sheet: str = "Jul 2026 PMS") -> dict:
    """Create Customer rows from the PMS sheet. Customer details only, deduped by CRM id."""
    ws = load_workbook(path, data_only=True, read_only=True)[sheet]
    rows = ws.iter_rows(min_row=2, values_only=True)

    # Existing dedupe indexes.
    existing_cid = {
        c.crm_customer_id for c in session.exec(select(Customer)).all() if c.crm_customer_id
    }
    existing_names = {
        c.name.lower() for c in session.exec(select(Customer)).all() if c.name
    }

    created = skipped = 0
    seen_cid: set[str] = set()
    seen_name: set[str] = set()

    for r in rows:
        # Column order (1-based): 4=Customer Id, 6=Account, 7=Address, 8=Zip, 9=City, 10=Phone
        cid = _s(r[3])
        name = _s(r[5])
        if not name:
            continue

        key = cid or f"name:{name.lower()}"
        # Skip duplicates within this file and against the DB.
        if cid and (cid in existing_cid or cid in seen_cid):
            skipped += 1
            continue
        if not cid and (name.lower() in existing_names or name.lower() in seen_name):
            skipped += 1
            continue

        session.add(Customer(
            name=name,
            crm_customer_id=cid,
            address=_s(r[6]),
            pincode=_s(r[7]),
            city=_s(r[8]),
            contact_number=_s(r[9]),
        ))
        created += 1
        seen_cid.add(cid) if cid else seen_name.add(name.lower())
        _ = key

    session.commit()
    return {"created": created, "skipped": skipped}


# ---------------------------------------------------------------------------
# Breakdown -> tickets
# ---------------------------------------------------------------------------
def _split_name_city(raw: str) -> tuple[str, str | None]:
    """'Dmart, Pollachi' -> ('Dmart', 'Pollachi'); no comma -> (raw, None)."""
    if "," in raw:
        name, city = raw.rsplit(",", 1)
        return name.strip(), city.strip() or None
    return raw.strip(), None


def _get_or_create_customer(session: Session, raw: str) -> Customer:
    name, city = _split_name_city(raw)
    found = session.exec(
        select(Customer).where(Customer.name.ilike(name))
    ).first()
    if found:
        return found
    cust = Customer(name=name, city=city)
    session.add(cust)
    session.commit()
    session.refresh(cust)
    return cust


def import_breakdown(session: Session, path: str, sheet: str = "Breakdown Status") -> dict:
    """Create Breakdown tickets from the sheet; auto-close rows with a Closing Date."""
    ws = load_workbook(path, data_only=True, read_only=True)[sheet]

    complaints = {c.name: c for c in session.exec(select(Complaint)).all()}

    created = skipped = closed = 0
    problems: list[str] = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        # 1-based: 2=Open Date, 3=Customer+Addr, 5=Machine, 7=Complaint, 8=Technician,
        #          10=Closing Date, 12=Remarks
        raw_cust = _s(r[2])
        open_date = _as_date(r[1])
        if not raw_cust or not open_date:
            continue

        raw_complaint = _s(r[6]) or "General Complaint"
        mapped_complaint = COMPLAINT_MAP.get(raw_complaint, raw_complaint)
        complaint = complaints.get(mapped_complaint)
        if complaint is None:
            problems.append(f"Unknown complaint '{raw_complaint}' (row skipped)")
            skipped += 1
            continue

        raw_machine = _s(r[4])
        machine = None
        if raw_machine:
            mval = MACHINE_MAP.get(raw_machine, raw_machine)
            try:
                machine = MachineType(mval)
            except ValueError:
                problems.append(f"Unknown machine '{raw_machine}' (row skipped)")
                skipped += 1
                continue

        cust = _get_or_create_customer(session, raw_cust)

        # Idempotency + "don't open a second ticket" rule.
        cust_tickets = session.exec(
            select(Ticket).where(Ticket.customer_id == cust.id)
        ).all()
        already = any(
            t.complaint_date == open_date and (t.skill or "").endswith(machine.value if machine else "")
            and _primary(t) == mapped_complaint
            for t in cust_tickets
        )
        if already:
            skipped += 1
            continue
        has_open = any(t.status != TicketStatus.CLOSED for t in cust_tickets)
        if has_open:
            skipped += 1
            continue

        # Create the ticket + Logged row (mirrors the create-ticket endpoint logic).
        skill = (
            derive_skill(complaint.complaint_type.value, machine)
            if machine else complaint.complaint_type.value
        )
        ticket = Ticket(
            ticket_no=next_ticket_no(session, WorkType.BREAKDOWN, open_date),
            customer_id=cust.id,
            complaint_date=open_date,
            work_type=WorkType.BREAKDOWN,
            machine_type=machine,
            skill=skill,
            status=TicketStatus.OPEN,
        )
        session.add(ticket)
        session.commit()
        session.refresh(ticket)

        session.add(TicketUpdate(
            ticket_id=ticket.id, stage=LifecycleStage.LOGGED, action_date=open_date,
            complaints=mapped_complaint, status=TicketStatus.OPEN,
        ))

        technician = _s(r[7])
        if technician:
            session.add(TicketUpdate(
                ticket_id=ticket.id, stage=LifecycleStage.ASSIGNED, action_date=open_date,
                job_lead=technician, complaints=mapped_complaint, status=TicketStatus.IN_PROGRESS,
            ))

        closing_date = _as_date(r[9])
        if closing_date:
            session.add(TicketUpdate(
                ticket_id=ticket.id, stage=LifecycleStage.CLOSED, action_date=closing_date,
                end_date=closing_date, remarks=_s(r[11]), status=TicketStatus.CLOSED,
            ))
            closed += 1

        session.commit()
        session.refresh(ticket)
        ticket.status = compute_ticket_status(ticket.updates)
        session.add(ticket)
        session.commit()
        created += 1

    return {"created": created, "skipped": skipped, "closed": closed, "problems": problems}


def _primary(ticket: Ticket) -> str | None:
    """The complaint recorded on the ticket's initial Logged row."""
    logged = next(
        (u for u in sorted(ticket.updates, key=lambda u: u.id or 0)
         if u.stage == LifecycleStage.LOGGED),
        None,
    )
    return logged.complaints if logged else None
