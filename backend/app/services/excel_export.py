"""Build .xlsx workbooks for data exports (openpyxl).

A "sheet" is a (title, headers, rows) tuple. Rows are lists of plain Python
values (str/int/float/date/None). Header row is styled and frozen, and column
widths are auto-sized within sensible bounds.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="334155")  # slate-700
_HEADER_ALIGN = Alignment(vertical="center")

Sheet = tuple[str, list[str], list[list]]


def _cell(value: object) -> object:
    """openpyxl handles date/number/str natively; coerce everything else to str."""
    if value is None or isinstance(value, (str, int, float, date)):
        return value
    return str(value)


def build_workbook(sheets: list[Sheet]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    for title, headers, rows in sheets:
        ws = wb.create_sheet(title[:31])  # Excel caps sheet names at 31 chars
        ws.append(headers)
        for c in ws[1]:
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = _HEADER_ALIGN
        for r in rows:
            ws.append([_cell(v) for v in r])
        ws.freeze_panes = "A2"

        for i, header in enumerate(headers, start=1):
            longest = len(str(header))
            for r in rows:
                if i - 1 < len(r) and r[i - 1] is not None:
                    longest = max(longest, len(str(r[i - 1])))
            ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 50)

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
